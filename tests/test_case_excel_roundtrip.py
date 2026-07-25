from openpyxl import Workbook, load_workbook

from core.case_import import import_case_histories_from_excel
from core.export_excel import (
    CASE_HISTORY_EXPORT_FIELDS,
    CASE_HISTORY_EXPORT_VERSION,
    export_case_histories_to_excel,
)
from db.case_repository import create_cases, list_cases


def _complete_case():
    return {
        "project": "Mine α",
        "domain": "Ore zone 1",
        "stope_id": "S-001",
        "surface": "Hanging wall",
        "depth_m": 425.5,
        "height_m": 30.0,
        "avg_dip_deg": 67.25,
        "width_m": 12.4,
        "span_m": 45.8,
        "q_prime": 4.2,
        "a": 0.8,
        "b": 0.7,
        "c": 5.5,
        "n": 12.936,
        "shape_factor_hr_m": 8.75,
        "stable_hr_limit_m": 7.1,
        "predicted_state": "Unstable",
        "calculation_mode": "Local",
        "standard_state": "Unstable",
        "local_state": "Stable",
        "local_boundary_name": "Site boundary",
        "local_boundary_n": 10.25,
        "actual_hr_m": 8.75,
        "observed_state": "Stable",
        "comment": "  Preserve this empirical note exactly  ",
    }


def _without_internal_id(row):
    return {key: value for key, value in row.items() if key != "id"}


def test_legacy_excel_format_still_imports(tmp_path):
    path = tmp_path / "legacy.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Mine", "Ore Zone", "Stope ID", "Surface", "N", "HR", "Fact", "Notes"])
    ws.append(["Old Mine", "D1", "10", "HW", "2,5", "4,75", "устойчиво", "legacy"])
    wb.save(path)

    rows = import_case_histories_from_excel(path)

    assert len(rows) == 1
    assert rows[0]["project"] == "Old Mine"
    assert rows[0]["surface"] == "Hanging wall"
    assert rows[0]["n"] == 2.5
    assert rows[0]["shape_factor_hr_m"] == 4.75
    assert rows[0]["observed_state"] == "Stable"


def test_full_export_import_preserves_every_case_field(tmp_path):
    source_db = tmp_path / "source.sqlite"
    workbook = tmp_path / "backup.xlsx"
    restored_db = tmp_path / "restored.sqlite"
    create_cases([_complete_case()], source_db)
    original = list_cases(source_db)

    export_case_histories_to_excel(original, workbook)
    imported = import_case_histories_from_excel(workbook)
    create_cases(imported, restored_db)
    restored = list_cases(restored_db)

    assert len(restored) == len(original) == 1
    assert _without_internal_id(restored[0]) == _without_internal_id(original[0])
    exported = load_workbook(workbook, data_only=True)
    assert exported["Metadata"]["B1"].value == CASE_HISTORY_EXPORT_VERSION
    assert exported["Case Histories"].max_column == len(CASE_HISTORY_EXPORT_FIELDS)


def test_full_export_tolerates_unknown_column(tmp_path):
    path = tmp_path / "future.xlsx"
    export_case_histories_to_excel([_complete_case()], path)
    wb = load_workbook(path)
    ws = wb["Case Histories"]
    ws.cell(1, ws.max_column + 1, "Future Engineering Field")
    ws.cell(2, ws.max_column, "future value")
    wb.save(path)

    rows = import_case_histories_from_excel(path)

    assert len(rows) == 1
    assert rows[0]["comment"] == _complete_case()["comment"]


def test_full_export_reports_missing_required_columns(tmp_path):
    path = tmp_path / "invalid.xlsx"
    export_case_histories_to_excel([_complete_case()], path)
    wb = load_workbook(path)
    ws = wb["Case Histories"]
    ws.delete_cols(1, 4)
    wb.save(path)

    try:
        import_case_histories_from_excel(path)
    except ValueError as error:
        assert "missing required columns" in str(error).lower()
    else:
        raise AssertionError("A clear validation error was expected")
