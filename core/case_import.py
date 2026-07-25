from pathlib import Path

from openpyxl import load_workbook

from db.case_repository import CASE_AUDIT_FIELDS, CASE_FIELDS


SUPPORTED_EXPORT_VERSIONS = {2}
FULL_EXPORT_SHEET = "Case Histories"


HEADER_ALIASES = {
    "project": [
        "project",
        "mine",
        "месторождение",
        "проект",
    ],
    "domain": [
        "domain",
        "ore zone",
        "ore_zone",
        "рудная зона",
        "рз",
        "домен",
    ],
    "stope_id": [
        "stope id",
        "stope_id",
        "stope",
        "camera",
        "камера",
        "номер камеры",
        "id камеры",
    ],
    "surface": [
        "surface",
        "face",
        "plane",
        "поверхность",
        "обнажение",
    ],
    "depth_m": [
        "depth",
        "depth m",
        "depth, m",
        "глубина",
        "глубина м",
        "глубина, м",
    ],
    "height_m": [
        "height",
        "height m",
        "height, m",
        "высота",
        "высота м",
        "высота, м",
        "высота этажа",
    ],
    "avg_dip_deg": [
        "avg dip",
        "average dip",
        "average dip, °",
        "average dip, deg",
        "average dip deg",
        "средний угол",
        "средний угол падения",
        "угол падения средний",
    ],
    "width_m": [
        "width",
        "width m",
        "width, m",
        "thickness",
        "ore thickness",
        "мощность",
        "ширина",
        "мощность, м",
    ],
    "span_m": [
        "span",
        "span m",
        "span, m",
        "strike length",
        "strike length, m",
        "пролет",
        "пролёт",
        "длина",
        "длина пролета",
        "длина пролёта",
        "пролет по простиранию",
        "пролёт по простиранию",
    ],
    "q_prime": [
        "q'",
        "q prime",
        "qprime",
        "q_prime",
        "q",
        "рейтинг бартона",
        "показатель рейтинга бартона",
    ],
    "a": [
        "a",
        "factor a",
        "stress factor",
        "stress factor a",
        "фактор a",
        "фактор напряжений",
    ],
    "b": [
        "b",
        "factor b",
        "joint factor",
        "joint orientation factor",
        "фактор b",
        "фактор ориентации трещин",
    ],
    "c": [
        "c",
        "factor c",
        "surface factor",
        "gravity factor",
        "фактор c",
        "гравитационный фактор",
    ],
    "n": [
        "n",
        "stability number",
        "mathews stability number",
        "показатель устойчивости",
        "показатель n",
    ],
    "shape_factor_hr_m": [
        "hr",
        "hydraulic radius",
        "hydraulic radius, m",
        "shape factor",
        "shape factor, m",
        "гидравлический радиус",
        "гидравлический радиус hr",
        "коэффициент формы",
    ],
    "stable_hr_limit_m": [
        "stable hr",
        "stable hr limit",
        "stable hr limit, m",
        "hr stable",
        "устойчивый hr",
    ],
    "predicted_state": [
        "predicted",
        "predicted state",
        "calculated state",
        "расчет",
        "расчёт",
        "расчетная оценка",
        "расчётная оценка",
    ],
    "observed_state": [
        "observed",
        "observed state",
        "actual state",
        "fact state",
        "fact",
        "факт",
        "фактическое состояние",
        "устойчивость",
        "категория устойчивости",
    ],
    "comment": [
        "comment",
        "comments",
        "note",
        "notes",
        "комментарий",
        "примечание",
    ],
    "calculation_mode": ["calculation mode", "calculation_mode"],
    "standard_state": ["standard state", "standard_state"],
    "local_state": ["local state", "local_state"],
    "local_boundary_name": ["local boundary name", "local boundary", "local_boundary_name"],
    "local_boundary_n": ["local boundary n", "boundary n", "local_boundary_n"],
    "actual_hr_m": ["actual hr", "actual hr, m", "actual_hr_m"],
    "created_at": ["created at", "created_at"],
    "updated_at": ["updated at", "updated_at"],
}


SURFACE_ALIASES = {
    "crown": "Crown",
    "back": "Crown",
    "roof": "Crown",
    "кровля": "Crown",
    "висячий бок": "Hanging wall",
    "hanging wall": "Hanging wall",
    "hw": "Hanging wall",
    "лежачий бок": "Footwall",
    "footwall": "Footwall",
    "fw": "Footwall",
    "end wall": "End wall",
    "endwall": "End wall",
    "борт": "End wall",
    "борта": "End wall",
    "торец": "End wall",
    "торцы": "End wall",
}


STATE_ALIASES = {
    "stable": "Stable",
    "устойчиво": "Stable",
    "устойчивая": "Stable",
    "устойчивые": "Stable",
    "устойчивый": "Stable",
    "un": "Unstable",
    "unstable": "Unstable",
    "failure": "Unstable",
    "minor failure": "Unstable",
    "неустойчиво": "Unstable",
    "неустойчивая": "Unstable",
    "неустойчивые": "Unstable",
    "неустойчивый": "Unstable",
    "caved": "Caved",
    "caving": "Caved",
    "major failure": "Caved",
    "обрушено": "Caved",
    "обрушенная": "Caved",
    "обрушенные": "Caved",
    "обрушенный": "Caved",
    "unknown": "Unknown",
    "": "Unknown",
}


def _normalize_text(value) -> str:
    if value is None:
        return ""

    text = str(value).strip().lower()
    text = text.replace("\n", " ")
    text = text.replace("ё", "е")

    while "  " in text:
        text = text.replace("  ", " ")

    return text


def _normalize_number(value):
    if value is None:
        return ""

    if isinstance(value, (int, float)):
        return value

    text = str(value).strip().replace(",", ".")

    if text == "":
        return ""

    try:
        return float(text)
    except ValueError:
        return str(value).strip()


def _normalize_surface(value) -> str:
    text = _normalize_text(value)

    return SURFACE_ALIASES.get(text, str(value).strip() if value is not None else "")


def _normalize_state(value) -> str:
    text = _normalize_text(value)

    return STATE_ALIASES.get(text, str(value).strip() if value is not None else "Unknown")


def _build_header_map(header_values: list) -> dict:
    normalized_headers = [_normalize_text(value) for value in header_values]

    header_map = {}

    for field, aliases in HEADER_ALIASES.items():
        normalized_aliases = [_normalize_text(alias) for alias in aliases]

        for column_index, header in enumerate(normalized_headers):
            if header in normalized_aliases:
                header_map[field] = column_index
                break

    return header_map


def _find_header_row(ws, max_rows: int = 20) -> tuple[int, dict]:
    for row_number in range(1, min(ws.max_row, max_rows) + 1):
        values = [cell.value for cell in ws[row_number]]
        header_map = _build_header_map(values)

        if "n" in header_map and "shape_factor_hr_m" in header_map:
            return row_number, header_map

    raise ValueError(
        "Could not find header row. Excel file must contain at least columns for N and HR."
    )


def _export_version(wb) -> int | None:
    if "Metadata" not in wb.sheetnames:
        return None

    ws = wb["Metadata"]
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
        if row and _normalize_text(row[0]) == "stopeforge export version":
            try:
                return int(row[1])
            except (IndexError, TypeError, ValueError) as error:
                raise ValueError("Invalid StopeForge Export Version in Metadata sheet.") from error
    return None


def _import_full_export(wb, version: int, sheet_name: str | None) -> list[dict]:
    if version not in SUPPORTED_EXPORT_VERSIONS:
        raise ValueError(
            f"Unsupported StopeForge Export Version: {version}. "
            f"Supported versions: {sorted(SUPPORTED_EXPORT_VERSIONS)}."
        )

    selected_sheet = sheet_name or FULL_EXPORT_SHEET
    if selected_sheet not in wb.sheetnames:
        raise ValueError(f"Data sheet '{selected_sheet}' was not found in workbook.")
    ws = wb[selected_sheet]
    header_map = _build_header_map([cell.value for cell in ws[1]])

    required = {"project", "domain", "stope_id", "surface"}
    missing = sorted(required - header_map.keys())
    if missing:
        raise ValueError("Full export is missing required columns: " + ", ".join(missing))

    defaults = {
        "project": "", "domain": "", "stope_id": "", "surface": "",
        "predicted_state": "", "calculation_mode": "Standard",
        "standard_state": "", "local_state": "", "local_boundary_name": "",
        "observed_state": "Unknown", "comment": "",
    }
    number_fields = {
        "depth_m", "height_m", "avg_dip_deg", "width_m", "span_m", "q_prime",
        "a", "b", "c", "n", "shape_factor_hr_m", "stable_hr_limit_m",
        "local_boundary_n", "actual_hr_m",
    }
    imported_rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in values):
            continue
        case_row = {}
        for field in [*CASE_FIELDS, *CASE_AUDIT_FIELDS]:
            column = header_map.get(field)
            value = values[column] if column is not None and column < len(values) else None
            if field in number_fields:
                case_row[field] = None if value in (None, "") else _normalize_number(value)
            elif value is None:
                case_row[field] = defaults.get(field, "")
            else:
                # Full exports already contain canonical application values.  Do not
                # rewrite them: comments and other user text must round-trip exactly.
                case_row[field] = str(value)
        imported_rows.append(case_row)
    return imported_rows


def import_case_histories_from_excel(path: str | Path, sheet_name: str | None = None) -> list[dict]:
    path = Path(path)

    wb = load_workbook(path, data_only=True)

    version = _export_version(wb)
    if version is not None:
        return _import_full_export(wb, version, sheet_name)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' was not found in workbook.")

        ws = wb[sheet_name]
    else:
        ws = wb.active

    header_row, header_map = _find_header_row(ws)

    imported_rows = []

    for row_number in range(header_row + 1, ws.max_row + 1):
        row_values = [cell.value for cell in ws[row_number]]

        raw_row = {}

        for field in CASE_FIELDS:
            column_index = header_map.get(field)

            if column_index is None or column_index >= len(row_values):
                raw_row[field] = ""
            else:
                raw_row[field] = row_values[column_index]

        if raw_row.get("n") in ("", None) and raw_row.get("shape_factor_hr_m") in ("", None):
            continue

        case_row = {field: "" for field in CASE_FIELDS}

        text_fields = [
            "project",
            "domain",
            "stope_id",
            "predicted_state",
            "comment",
        ]

        number_fields = [
            "depth_m",
            "height_m",
            "avg_dip_deg",
            "width_m",
            "span_m",
            "q_prime",
            "a",
            "b",
            "c",
            "n",
            "shape_factor_hr_m",
            "stable_hr_limit_m",
        ]

        for field in text_fields:
            value = raw_row.get(field, "")
            case_row[field] = "" if value is None else str(value).strip()

        for field in number_fields:
            case_row[field] = _normalize_number(raw_row.get(field, ""))

        case_row["surface"] = _normalize_surface(raw_row.get("surface", ""))
        case_row["observed_state"] = _normalize_state(raw_row.get("observed_state", "Unknown"))

        if case_row["predicted_state"]:
            case_row["predicted_state"] = _normalize_state(case_row["predicted_state"])

        if case_row["project"] == "":
            case_row["project"] = path.stem

        if case_row["observed_state"] == "":
            case_row["observed_state"] = "Unknown"

        imported_rows.append(case_row)

    return imported_rows
