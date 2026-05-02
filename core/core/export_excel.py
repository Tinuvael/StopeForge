from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.models import StopeResult


THIN_BORDER = Border(
    left=Side(style="thin", color="999999"),
    right=Side(style="thin", color="999999"),
    top=Side(style="thin", color="999999"),
    bottom=Side(style="thin", color="999999"),
)

TITLE_FILL = PatternFill("solid", fgColor="D9EAD3")
SECTION_FILL = PatternFill("solid", fgColor="E2F0D9")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
RESULT_FILL = PatternFill("solid", fgColor="D9EAF7")
STABLE_FILL = PatternFill("solid", fgColor="A9D18E")
UNSTABLE_FILL = PatternFill("solid", fgColor="F4B183")
CAVED_FILL = PatternFill("solid", fgColor="E06666")
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")


def _style_range(ws, cell_range: str, fill=None, bold=False, align_center=True):
    for row in ws[cell_range]:
        for cell in row:
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            if bold:
                cell.font = Font(name="Times New Roman", size=11, bold=True)
            else:
                cell.font = Font(name="Times New Roman", size=11)

            if align_center:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _merge_title(ws, cell_range: str, text: str):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = text
    cell.font = Font(name="Times New Roman", size=12, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = TITLE_FILL
    _style_range(ws, cell_range, fill=TITLE_FILL, bold=True)


def _state_fill(state: str):
    state_lower = state.lower()

    if "stable" in state_lower and "un" not in state_lower:
        return STABLE_FILL

    if "unstable" in state_lower:
        return UNSTABLE_FILL

    if "caved" in state_lower:
        return CAVED_FILL

    return RESULT_FILL


def _safe_value(value: Any):
    if value == float("inf"):
        return "not limited"

    return value


def _fmt(value: float, digits: int = 2):
    if value == float("inf"):
        return "not limited"

    return round(value, digits)


def export_current_calculation_to_excel(
    result: StopeResult,
    joint_sets: list,
    output_path: str | Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Calculation"

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16

    for row in range(1, 45):
        ws.row_dimensions[row].height = 22

    stope = result.stope

    # ------------------------------------------------------------------
    # Input data / Rock mass
    # ------------------------------------------------------------------
    _merge_title(ws, "A1:G1", "Input Data")
    _merge_title(ws, "A2:G2", "Rock Mass Characteristics")

    rock_rows = [
        ("H, Mining depth", "m", stope.depth_m),
        ("ρ, Unit weight", "t/m³", stope.unit_weight_t_m3),
        ("λ, Horizontal stress ratio", "-", stope.horizontal_stress_ratio),
        ("σc, UCS", "MPa", stope.ucs_mpa),
    ]

    start_row = 3
    for i, (name, unit, value) in enumerate(rock_rows, start=start_row):
        ws[f"A{i}"] = name
        ws[f"D{i}"] = unit
        ws[f"G{i}"] = value

    _style_range(ws, f"A{start_row}:G{start_row + len(rock_rows) - 1}", align_center=False)
    _style_range(ws, f"G{start_row}:G{start_row + len(rock_rows) - 1}", fill=INPUT_FILL)

    # ------------------------------------------------------------------
    # Joint sets
    # ------------------------------------------------------------------
    joint_title_row = 7
    _merge_title(ws, f"A{joint_title_row}:G{joint_title_row}", "Main Discontinuity Set Orientations")

    ws["A8"] = ""
    for idx in range(1, 6):
        ws.cell(row=8, column=idx + 2).value = f"Set {idx}"

    ws["A9"] = "Dip angle"
    ws["B9"] = "deg"
    ws["A10"] = "Dip direction"
    ws["B10"] = "deg"

    for idx, joint in enumerate(joint_sets[:5], start=3):
        ws.cell(row=9, column=idx).value = joint.dip_deg
        ws.cell(row=10, column=idx).value = joint.dip_direction_deg

    _style_range(ws, "A8:G10")
    _style_range(ws, "C9:G10", fill=INPUT_FILL)

    # ------------------------------------------------------------------
    # Stope geometry
    # ------------------------------------------------------------------
    stope_title_row = 11
    _merge_title(ws, f"A{stope_title_row}:G{stope_title_row}", "Stope Geometry")

    stope_rows = [
        ("Stope height", "m", stope.stope_height_m),
        ("Average stope dip", "deg", stope.average_dip_deg),
        ("Stope span / strike length", "m", stope.stope_span_m),
        ("Stope width / ore thickness", "m", stope.stope_width_m),
        ("Hanging wall dip direction", "deg", stope.hanging_wall_dip_direction_deg),
    ]

    start_row = 12
    for i, (name, unit, value) in enumerate(stope_rows, start=start_row):
        ws[f"A{i}"] = name
        ws[f"D{i}"] = unit
        ws[f"G{i}"] = value

    _style_range(ws, f"A{start_row}:G{start_row + len(stope_rows) - 1}", align_center=False)
    _style_range(ws, f"G{start_row}:G{start_row + len(stope_rows) - 1}", fill=INPUT_FILL)

    # ------------------------------------------------------------------
    # Stability assessment
    # ------------------------------------------------------------------
    eval_title_row = 17
    _merge_title(ws, f"A{eval_title_row}:G{eval_title_row}", "Stability Assessment Parameters")

    # Header row
    ws.merge_cells("A18:C18")
    ws["A18"] = "Parameter"
    ws["D18"] = "Crown"
    ws["E18"] = "Hanging wall"
    ws["F18"] = "Footwall"
    ws["G18"] = "End wall"

    _style_range(ws, "A18:G18", fill=HEADER_FILL, bold=True)

    surface_map = {surface.surface_type.value: surface for surface in result.surfaces}
    ordered_surface_names = ["Crown", "Hanging wall", "Footwall", "End wall"]

    parameter_rows = [
        ("Q', Barton rating", "q_prime"),
        ("A, stress factor", "stress_factor_a"),
        ("B, joint orientation factor", "joint_factor_b"),
        ("C, surface orientation factor", "surface_factor_c"),
        ("N, stability number", "stability_number_n"),
        ("Actual HR", "actual_hr_m"),
        ("HR, stable hydraulic radius", "hr_stable"),
        ("HR, caving hydraulic radius", "hr_caving"),
        ("Standard assessment", "stability_state"),
        ("Local assessment", "local_state"),
        ("Local boundary", "local_boundary_name"),
        ("Boundary N", "local_boundary_n"),
    ]


    first_data_row = 19

    for row_idx, (label, attr) in enumerate(parameter_rows, start=first_data_row):
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
        ws.cell(row=row_idx, column=1).value = label

        for col_idx, surface_name in enumerate(ordered_surface_names, start=4):
            surface = surface_map.get(surface_name)

            if surface is None:
                ws.cell(row=row_idx, column=col_idx).value = None
                continue

            value = getattr(surface, attr, None)

            if attr in ("stability_state", "local_state"):
                value = "" if value is None else value.value
            elif isinstance(value, float):
                value = _fmt(value, 2)
            elif value is None:
                value = ""

            ws.cell(row=row_idx, column=col_idx).value = _safe_value(value)


    last_assessment_row = first_data_row + len(parameter_rows) - 1

    _style_range(ws, f"A{first_data_row}:G{last_assessment_row}")

    # Style merged parameter cells manually
    for row_idx in range(first_data_row, last_assessment_row + 1):
        for col_idx in range(1, 4):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.font = Font(name="Times New Roman", size=11)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Color Standard / Local assessment rows
    for row_idx, (label, attr) in enumerate(parameter_rows, start=first_data_row):
        if attr not in ("stability_state", "local_state"):
            continue

        for col in range(4, 8):
            cell = ws.cell(row=row_idx, column=col)

            if cell.value:
                cell.fill = _state_fill(str(cell.value))
                cell.font = Font(name="Times New Roman", size=11, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")


    # ------------------------------------------------------------------
    # Summary
    # ------------------------------------------------------------------
    summary_row = last_assessment_row + 1
    _merge_title(ws, f"A{summary_row}:G{summary_row}", "Summary")

    calculation_mode = getattr(result, "calculation_mode", "Standard")
    local_final_state = getattr(result, "local_final_state", None)

    summary_items = [
        ("Calculation mode", calculation_mode),
        ("Standard final state", result.final_state.value),
        (
            "Local final state",
            "" if local_final_state is None else local_final_state.value,
        ),
        ("Limiting surface", result.limiting_surface.value),
    ]

    for offset, (label, value) in enumerate(summary_items, start=1):
        row_idx = summary_row + offset

        ws[f"A{row_idx}"] = label
        ws.merge_cells(
            start_row=row_idx,
            start_column=2,
            end_row=row_idx,
            end_column=3,
        )
        ws[f"B{row_idx}"] = value

    last_summary_row = summary_row + len(summary_items)

    _style_range(ws, f"A{summary_row + 1}:G{last_summary_row}", align_center=False)

    # Color final state cells
    ws[f"B{summary_row + 2}"].fill = _state_fill(result.final_state.value)
    ws[f"B{summary_row + 2}"].font = Font(name="Times New Roman", size=11, bold=True)
    ws[f"B{summary_row + 2}"].alignment = Alignment(horizontal="center", vertical="center")

    if local_final_state is not None:
        ws[f"B{summary_row + 3}"].fill = _state_fill(local_final_state.value)
        ws[f"B{summary_row + 3}"].font = Font(name="Times New Roman", size=11, bold=True)
        ws[f"B{summary_row + 3}"].alignment = Alignment(horizontal="center", vertical="center")


    ws.freeze_panes = "A18"
    ws.sheet_view.showGridLines = False

    wb.save(output_path)



def export_project_overview_to_excel(
    rows: list[dict],
    output_path: str | Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Overview"

    headers = [
        "Surface",
        "Dip, °",
        "Q'",
        "A",
        "B",
        "C",
        "N",
        "Actual HR",
        "HR stable",
        "HR caving",
        "Stable span",
        "Caving span",
        "Rating length",
        "Standard State",
        "Local State",
        "Local Boundary",
        "Boundary N",
    ]


    ws.append(headers)

    for row in rows:
        ws.append(
            [
                row.get("project", ""),
                row.get("domain", ""),
                row.get("stope_id", ""),
                row.get("depth", ""),
                row.get("height", ""),
                row.get("avg_dip", ""),
                row.get("width", ""),
                row.get("span", ""),
                row.get("limiting_surface", ""),
                row.get("final_state", ""),
                row.get("comment", ""),
            ]
        )

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(name="Times New Roman", size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Times New Roman", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

            if cell.column == 10:
                cell.fill = _state_fill(str(cell.value))

    widths = {
        "A": 18,
        "B": 18,
        "C": 16,
        "D": 12,
        "E": 12,
        "F": 16,
        "G": 12,
        "H": 12,
        "I": 20,
        "J": 18,
        "K": 35,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    wb.save(output_path)
